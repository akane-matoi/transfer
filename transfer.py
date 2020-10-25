from selenium import webdriver         # Webブラウザを自動操作する（python -m pip install selenium)
from selenium.webdriver.common.keys import Keys # webdriverからスクレイピングで使用するキーを使えるようにする。
from selenium.webdriver.chrome.options import Options
import chromedriver_binary
from selenium.webdriver.support.ui import Select
import os
import sys
import time
import openpyxl
import datetime

now = datetime.datetime.now()

wb = openpyxl.load_workbook('list.xlsx',data_only=True)
login = openpyxl.load_workbook('user.xlsx',data_only=True)

sheet = wb["Sheet1"]
sheet2 = login["sheet2"]

print("自動拠点間移動を開始します。")
print("salesforceへのログイン情報確認")

user = sheet2.cell(row=2,column=2).value
fpass = sheet2.cell(row=3,column=2).value


if not user:
    print('ユーザー名が未入力です')
    sys.exit()
elif not fpass:
    print('パスワードが未入力です')
    sys.exit()
else:
    print("ユーザーID:"+str(user))
    print("パスワード:"+str(fpass))

print('よろしいですか?y/n')
select = input()

if select== "n":
    print('終了します。')
    sys.exit()
else:
    print()

b = 6
c = 7
note2 = sheet.cell(row = 4,column = 2).value#←エクセルから読み取り
times = now.strftime('20%y/%m/%d')
#times2 = now.strftime('20%y/%m/%d')
time2 = sheet.cell(row=5,column=2).value
base = sheet.cell(row=2,column=2).value
qty = sheet.cell(row = 1,column = 2).value

if not time2:
    print('到着予定日が未入力です')
    sys.exit()
elif not qty:
    print('P/N,S/Nが未入力です')
    sys.exit()
else:
    print()

print('以下内容で移動開始します。')
print(note2)

print('実行しますか?y/n')
select = input()

if select== "n":
    print('終了します。')
    sys.exit()
else:
    print()

time.sleep(3)

#---------------------------------

#options = webdriver.ChromeOptions()
#options.add_argument('--user-data-dir=C:UsersUserNameAppDataLocalGoogleChromeUser Data')
#options.add_argument('--profile-directory=Default')  # この行を省略するとDefaultフォルダが指定されます
userdata_dir = 'UserData'
os.makedirs(userdata_dir, exist_ok=True)

options = webdriver.ChromeOptions()
options.add_argument('--user-data-dir=' + userdata_dir)
driver = webdriver.Chrome(options=options)

driver.get('https://login.salesforce.com/')  # salesforceを開く
input = driver.find_element_by_id('username')#ユーザー入力
input.send_keys(user)
input = driver.find_element_by_id('password')#パスワード入力
input.send_keys(fpass)
input = driver.find_element_by_id('Login').click()

time.sleep(3)

while qty > 0 :
    qty = qty - 2
    pn = sheet.cell(row = b,column = 2).value#←エクセルから読み取り
    sn = sheet.cell(row = c,column = 2).value#←エクセルから読み取り

    driver.get('https://ap3.salesforce.com/a02/e?retURL=%2Fa02%2Fo')
    time.sleep(3)

    Type = driver.find_element_by_name("00N10000003g0aj")
    reason_select = Select(Type)
    reason_select.select_by_value("Ship only")

    input = driver.find_element_by_id('CF00N10000002QnJm')#P/N
    input.send_keys(pn)

    input = driver.find_element_by_id('00N10000002QnK6')#QTY
    input.send_keys(1)

    input = driver.find_element_by_id('00N10000002QnKB')#S/N
    input.send_keys(sn)

    input = driver.find_element_by_id('00N10000002QnJw')
    input.send_keys(times)

    input = driver.find_element_by_id('00N10000002SLNX')#NOte2
    input.send_keys(note2)

    list = driver.find_element_by_name("00N10000002Xu2o")
    list_select = Select(list)
    list_select.select_by_value("ITSS 東京")

    selector = '#topButtonRow > input:nth-child(1)'
    element = driver.find_element_by_css_selector(selector)
    driver.execute_script('arguments[0].click();', element)

    time.sleep(3)

    driver.get('https://ap3.salesforce.com/a00/e?retURL=%2Fa00%2Fo')

    reason = driver.find_element_by_name("00N10000003ccSh")
    reason_select = Select(reason)
    reason_select.select_by_value("New")

    input = driver.find_element_by_id('CF00N10000002YqNJ')#P/N
    input.send_keys(pn)

    input = driver.find_element_by_id('00N10000002YqNO')#QTY
    input.send_keys(1)

    input = driver.find_element_by_id('00N10000002YqNT')#S/N
    input.send_keys(sn)

    input = driver.find_element_by_id('00N10000002YqNs')
    input.send_keys(time2)

    input = driver.find_element_by_id('00N10000002YqOR')#NOte2
    input.send_keys(note2)

    list = driver.find_element_by_name("00N10000002YqNn")
    list_select = Select(list)
    list_select.select_by_value(base)

    selector = '#topButtonRow > input:nth-child(1)'
    element = driver.find_element_by_css_selector(selector)
    driver.execute_script('arguments[0].click();', element)

    c = c + 2
    b = b + 2

print('入庫終了します。')
driver.quit()
sys.exit()
